# _*_ coding:utf-8 _*_
import random
# from translate import translate
from openpyxl import load_workbook

class WordModel:
    """
    word指单词本身
    meanings 指单词释义
    score指单词得分情况
    """
    def __init__(self,row = 0,word = '' ,sheet = None , meaning = '',father = None,listid = -1,correct = 0,false = 0,col_num=0):
        self.row = row
        self.word = word
        self.meaning = meaning
        self.sheet = sheet
        self.father = father
        self.listid = listid
        self.correct = correct
        self.false = false
        self.col_num = col_num
class Controller:
    """

    """

    rowlist = []
    sheetlist = []

    def __init__(self,username = ''):
        self.username = username
        self.wordlist = []

    def input_name(self,name):
        self.username = name
    def output_name(self):
        return self.username

    def add_word(self,word):
        self.wordlist.append(word)

    @property
    def get_word_list(self):
        return self.wordlist

    def get_all_meaning(self,i):
        father = i.father
        if father:
            for wordrow in range(father.listid,self.wordlist[-1].listid):
                if self.wordlist[wordrow].word == father.word:
                    print(self.wordlist[wordrow].meaning)
                else:
                    break
        else:print('这个单词没有提示。。。')


    def randomize_testsheet(self,sheet,command =0):
        list_for_random = []
        if command == 0:
            #随机抽取
            if sheet != 0:
                flag = 0
                for i in self.wordlist:
                    if i.sheet == sheet:
                        list_for_random.append(i)
                        flag = 1
                    elif i.sheet != sheet and flag == 1:
                        break
                    elif i.sheet != sheet and flag == 0:
                        pass
            else:
                for i in self.wordlist:
                    list_for_random.append(i)
            random.shuffle(list_for_random)
            return list_for_random
        elif command == 3:
            #顺序抽取
            if sheet != 0:
                flag = 0
                for i in self.wordlist:
                    if i.sheet == sheet:
                        list_for_random.append(i)
                        flag = 1
                    elif i.sheet != sheet and flag == 1:
                        break
                    elif i.sheet != sheet and flag == 0:
                        pass
            else:
                for i in self.wordlist:
                    list_for_random.append(i)
            return list_for_random
        elif command == 1:
            if sheet != 0:
                flag = 0
                for i in self.wordlist:
                    if i.sheet == sheet:
                        list_for_random.append(i)
                        flag = 1
                    elif i.sheet != sheet and flag == 1:
                        break
                    elif i.sheet != sheet and flag == 0:
                        pass
            else:
                for i in self.wordlist:
                    list_for_random.append(i)
            list_for_random.sort(key = lambda i: (i.correct)/(i.correct+i.false))
            return list_for_random



    def choice_appender(self,word,testlist,choices = 4):
        choice_list = []
        choice_list.append(word)
        while len(choice_list) < choices:
            i = testlist[random.randint(1,len(testlist))-1]
            if i.word != word.word and not i in choice_list:
                choice_list.append(i)
            else:
                pass
        random.shuffle(choice_list)
        return choice_list


class View:
    def __init__(self):
        self.__controller = Controller()

    def main(self):
        self.importer()
        while True:
            self.print_menu()
            self.__select_menu_item()
            # os.system("cls")

    def importer(self):
        for sheet in sheetlist:
            row = 2
            word_colume = element_locater('#带井号的不会统计', sheet)
            meanings_colume = element_locater('#解释', sheet)
            while blank_checker(sheet,row):
                value_word = sheet[word_colume+ str(row)].value

                # print(value_word)
                #测试

                if value_word == None:
                    word = WordModel()
                    word.word = self.__controller.get_word_list[-1].word
                    word.row = row
                    word.meaning = sheet[meanings_colume + str(row)].value
                    word.sheet = sheet
                    word.father = self.__controller.get_word_list[-1]
                    word.listid = len(self.__controller.get_word_list)
                    self.__controller.add_word(word)
                elif value_word[0] ==  '#':
                    pass
                else:
                    word= WordModel()
                    word.word = value_word
                    word.row = row
                    word.meaning = sheet[meanings_colume + str(row)].value
                    word.sheet = sheet
                    word.listid = len(self.__controller.get_word_list)
                    self.__controller.add_word(word)

                row += 1

    def print_menu(self):
        print("1)登陆或注册")
        print("2)测试")
        print("3)学习")
        print("4)显示所有单词")
        print("5)退出")
        # print(self.__controller.wordlist[2].word)

    def __select_menu_item(self):
        item = input("请您输入选项:")
        if item == "1":
            if self.checker:
                print("已经登陆，若要退出，请重新加载程序")
            else:
                self.login()
        elif item == "2":
            self.test()
        elif item == "3":
            self.learn()
        elif item == "4":
            self.show_word()
        elif item == "5":
            forms.save("words.xlsx")
            forms.close()
            exit(0)
        else:
            print("输入错误")

    def username_login_sheet(self,command = 'f',target_sheet=None):
        #command 用于标记操作 默认是‘f’表示find也就是查找用户在某个单词表里是否存在
        #target_sheet 
        #所有command:
        #   1.'f'   find 查找用户是否存在 存在print输出 无返回值
        #   2.'c'   create 在指定单词表里创建用户
        #   3.'t'   test 程序内部判断是否存在用户，没有输出和创建请求
        name = self.__controller.output_name()
        # wordlist = self.__controller.get_word_list()
        namecolume = element_locater(name , target_sheet ,row = '2')
        if command == 'f':
            if namecolume:
                print("在" ,end='  ')
                print(target_sheet,end= '  ')
                print("中找到用户:  "+name)
                self.correct_rate_importer(target_sheet)
                return True
            else:
                print("在" ,end='  ')
                print(target_sheet,end= '  ')
                print("中没有找到用户:  ",name)
                req = input("是否创建？ y or n")
                if req == 'y':
                    self.username_login_sheet('c',target_sheet)
                    return True
                elif req == 'n':
                    return False
        elif command == 'c':
            namecolume = element_locater("user_end_point",target_sheet,row='2')
            target_sheet.insert_cols(target_sheet[namecolume+'2'].col_idx)
            target_sheet[namecolume+'2'].value = name
            print("在",target_sheet,"中创建用户成功")
            self.correct_rate_importer(target_sheet)
        elif command == 't':
            if namecolume:
                return True
            else:return False


    def login(self):

        name = input("输入用户名")
        self.__controller.input_name(name)
        # wordlist = self.__controller.get_word_list
        # current_sheet = wordlist[1].sheet
        # namecolume = element_locater(name , current_sheet ,row = '2')
        # if namecolume:
            # print("在" ,end='  ')
            # print(current_sheet,end= '  ')
            # print("中找到用户:  "+name)
            # self.correct_rate_importer(current_sheet)
        



    def correct_rate_importer(self,target_sheet):
        name = self.__controller.output_name()
        word_colume = element_locater('#带井号的不会统计', target_sheet)
        flag = 0
        if name and word_colume:
            # print('correctname importer test:',name)
            
            for word in self.__controller.wordlist:
                if word.sheet == target_sheet:
                    flag = 1
                    namecolume = element_locater(name , target_sheet , row = '2')
                    word.col_num = target_sheet[namecolume+'2'].col_idx
                    rate_chr = target_sheet[namecolume+str(word.row)].value
                        
                    if rate_chr == None:
                        word.correct = 0
                        word.false = 0
                    else:
                        correct = ''
                        all_ = ''
                        slice_pos = rate_chr.index('、')
                        for i in range(slice_pos):
                            correct+= rate_chr[i]
                        word.correct = int(correct)
                            # print(correct)
                        for i in range(slice_pos+1,len(rate_chr)):
                            all_+= rate_chr[i]
                        word.false = int(all_)-int(correct)
                elif flag == 1:break
                elif flag == 0:pass
            else: return False
        else:return False
    @property
    def select_sheet(self):
        flag = 0
        for mn in range(len(sheetlist)):
            flag+=1
            print(str(flag)+ " >> ", sheetlist[mn])
        ss = input("选择第几个表格或者输入a全选")
        if ss == 'a':
            return 0
        else:
            return sheetlist[int(ss)-1]

    @property
    def checker(self):
        if self.__controller.output_name() == '':
            print("请先登陆")
            return False
        else:
            return True

    def test(self):
        print('============================')
        if self.checker:
            i = self.select_sheet
            if self.username_login_sheet('f',i):
                m = int(input("输入0来随机抽取 1来按照正确率抽取(低到高) 2来按照正确率反向抽取（高到低） 3来顺序抽取"))
                testlist = self.__controller.randomize_testsheet(i,m)
                if input("输入1来根据单词选词义，2来根据词义默写单词") == '1':
                    self.test_by_choose_meaning(testlist)
                else:self.test_by_enter_word(testlist)
            else:print("选中单词表不存在用户，请创建或者更换用户")
            

    def test_by_choose_meaning(self,testlist,repeat = True):
        print("注意一定要大写字母选项，只能填 ABCD等等   输入0来退出")
        for i in testlist:
            print(i.word)
            chioce_list = self.__controller.choice_appender(i, testlist, 4)
            for m in range(len(chioce_list)):
                print("%s:%s" % (chr(65 + m), chioce_list[m].meaning))
            chioce = input("请选择：")
            if chioce == '0':
                break
            chioce = ord(chioce) - 65
            if chioce_list[chioce].word == i.word or chioce_list[chioce].meaning == i.meaning:
                i.correct+=1
                print('row',i.row)
                print('col',i.col_num)
                cell =i.sheet.cell(i.row , i.col_num)  #先行再列，这个东西和别的反过来的
                cell.value = correct_rate_adder('r',cell.value)
                print('=========================')
                print("恭喜你答对了")
                print('word:      ', i.word)
                print('meanings:  ', i.meaning)
                print("=========================")
            else:
                i.false+=1
                cell =i.sheet.cell(i.row , i.col_num)  #先行再列，这个东西和别的反过来的
                cell.value = correct_rate_adder('f',cell.value)
                print("=========================")
                print("答错了。。。正确答案为：%s" % (i.meaning))
                print("=========================")
                if repeat : testlist.append(i)

    def test_by_enter_word(self,testlist,repeat = True):
        for i in testlist:
            print(i.meaning)
            answer = input("请输入单词，输入1可以获取提示，输入0退出")
            if answer == i.word:
                print('=========================')
                print("恭喜你答对了")
                print('word:      ', i.word)
                print('meanings:  ', i.meaning)
                print("=========================")
            elif answer == '0':
                break
            elif answer =='1':
                self.__controller.get_all_meaning(i)
                answer = input("现在呢？会了吗？手动滑稽")
                if answer == i.word:
                    print('=========================')
                    print("恭喜你答对了")
                    print('word:      ', i.word)
                    print('meanings:  ', i.meaning)
                    print("=========================")
                else:
                    print("=========================")
                    print("答错了。。。正确答案为：%s" % (i.word))
                    print("=========================")
                    if repeat: testlist.append(i)
            else:
                print("=========================")
                print("答错了。。。正确答案为：%s" % (i.word))
                print("=========================")
                if repeat : testlist.append(i)


    def learn(self):
        if self.checker:
            i = self.select_sheet
            if i == 0:
                pass
            else:
                pass

    def show_word(self):
        wordlist = self.__controller.get_word_list
        for i in wordlist:
            print('word:          ' ,  i.word)
            print('meanings:      ' , i.meaning)
            print('sheet:         ' ,  i.sheet)
            print('row:           ' ,  i.row)
            if self.username_login_sheet('t',i.sheet):
                self.correct_rate_importer(i.sheet)
                # print("coorect rate test",i.correct)
                if i.correct+i.false != 0 : print('correct_rate:  ' ,  i.correct/(i.correct+i.false))
            print('=========================')

userlist = []
sheetlist = []
print("initializing........")
# localpath = os.getcwd()
# dirinformation = os.listdir(localpath)
# print(dirinformation)
forms = load_workbook("words.xlsx")
print("找到单词本：")

def initialize_user(sheet,listnumber):
    for ascid in range(ord('A'),ord('Z')):
        cellname = chr(ascid)+'1'
        # print(cellname)
        # print(sheet[cellname].value)
        if sheet[cellname].value == '#单个人':
            userlist[listnumber][sheets].append(ascid)
            for b in range(ascid,ord('Z')):
                cellname_2 = chr(b)+'2'
                if sheet[cellname_2].value != None:
                    userlist[listnumber][sheets].append(sheet[cellname_2].value)
                else:
                    break
            break

def correct_rate_adder(right_or_not,rate_chr):
    if not rate_chr == None:
        correct = ''
        all_ = ''
        slice_pos = rate_chr.index('、')
        for i in range(slice_pos):
            correct+= rate_chr[i]
        
        for i in range(slice_pos+1,len(rate_chr)):
            all_+= rate_chr[i]
        
        if right_or_not == 'r':
            correct = str(int(correct)+1)
            all_ = str(int(all_)+1)
        elif right_or_not =='f':
            all_ = str(int(all_)+1)
        return correct+'、'+all_
    else:
        if right_or_not == 'r': return '1、1'
        elif right_or_not == 'f': return '0、1'
        


def element_locater(element,sheet,row = '1'):
    flag = False
    for ascid in range(ord('A'),ord('Z')):
        cellname = chr(ascid)+ row
        # print(cellname)
        # print(sheet[cellname].value)
        if sheet[cellname].value == element:
            flag = True
            return chr(ascid)
    return flag

def blank_checker(sheet,row):
    bo = True
    for ascid in range(ord('A'), ord('F')):
        cellname = chr(ascid) + str(row)
        if sheet[cellname].value != None:
            return True
        else:
            bo = False
    return bo
a = 0
for sheets in forms:
    sheetlist.append(sheets)
    userlist.append({})
    userlist[a][sheets]=[]
    print(sheets)
    initialize_user(sheets,a)
    # print(userlist)
    a+=1




# well = translate()
# well.getword("well")
# print(well.translation())
#测试translate
View().main()
